
# importing whole module 
from tkinter import * 
from tkinter.ttk import *
from openpyxl import *
from time import strftime

wb = load_workbook('Stunden.xlsx')
ws = wb.worksheets[0]
i = 0
info = []

def get_data():
    x = 2
    cells = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    tcell = ws['A2'].value
    while tcell != None:
        cell_A = ws[cells[0] + str(x)].value
        cell_B = ws[cells[1] + str(x)].value
        cell_C = ws[cells[2] + str(x)].value
        cell_D = ws[cells[3] + str(x)].value
        cell_E = ws[cells[4] + str(x)].value
        cell_F = ws[cells[5] + str(x)].value
        cell_G = ws[cells[6] + str(x)].value
        data = [cell_A, cell_B, cell_C, cell_D, cell_E, cell_F, cell_G]
        info.append(data)
        data = []
        x += 1
        tcell = ws[cells[0] + str(x)].value

def kommen():
    get_data()
    ws[info[1]].value = "1"
    kommen = 1
    gehen = 0
    
    
    print (kommen, gehen)
    
def gehen():
    kommen = 0
    gehen = 1 
    print (kommen, gehen)

wb.save(filename = 'Stunden.xlsx')  
# creating tkinter window 
root = Tk() 
root.title('Clock') 
  
# This function is used to  
# display time on the label 
def time(): 
    string = strftime('%H:%M:%S') 
    lbl.config(text = string) 
    lbl.after(1000, time) 
  
# Styling the label widget so that clock 
# will look more attractive 
lbl = Label(root, font = ('calibri', 40, 'bold'), 
            background = 'purple', 
            foreground = 'white') 
kommen = Button(root, text = "Kommen", command = kommen)
kommen.pack(anchor = 'center')
gehen = Button(root, text = "Gehen", command = gehen)
gehen.pack(anchor = 'center')
  
# Placing clock at the centre 
# of the tkinter window 
lbl.pack(anchor = 's') 
time() 
  
mainloop() 
